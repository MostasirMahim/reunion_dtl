"""
Excel export for the admin dashboard.

Builds a formatted .xlsx workbook from a Registrant queryset — styled header,
frozen panes, auto-filter, sized columns, banded rows, colour-coded payment
status and a small summary block at the top.

Requires openpyxl (see requirements.txt).
"""

from django.conf import settings
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --- palette (matches the site's navy / gold identity) ---------------------
NAVY = "0B2D57"
NAVY_DARK = "071D3A"
GOLD = "C79A3C"
CREAM = "FAF6EE"
BAND = "F4F7FB"
WHITE = "FFFFFF"
GREY_TEXT = "5B6B80"

STATUS_FILL = {
    "paid": ("DCFCE7", "166534"),
    "pending": ("FEF3C7", "92400E"),
    "failed": ("FEE2E2", "991B1B"),
    "cancelled": ("E5E7EB", "374151"),
}

THIN = Side(style="thin", color="D8E0EA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("#", 6),
    ("Registration ID", 18),
    ("Full Name", 26),
    ("Phone", 15),
    ("Secondary Phone", 16),
    ("WhatsApp", 15),
    ("Email", 28),
    ("Last Class Attended", 20),
    ("SSC Batch", 11),
    ("SSC Passing Year", 15),
    ("Blood Group", 12),
    ("T-Shirt Size", 12),
    ("Present Address", 32),
    ("Amount (BDT)", 13),
    ("Payment Status", 15),
    ("Transaction ID", 24),
    ("Paid At", 18),
    ("Checked In", 11),
    ("Checked In At", 18),
    ("Registered At", 18),
]


def _local(dt):
    if not dt:
        return ""
    return timezone.localtime(dt).strftime("%d %b %Y, %I:%M %p")


def build_registrants_workbook(queryset, filter_summary=""):
    """Return an openpyxl Workbook for the given Registrant queryset."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"

    last_col = len(COLUMNS)
    last_letter = get_column_letter(last_col)

    # ---- title block -----------------------------------------------------
    ws.merge_cells(f"A1:{last_letter}1")
    title = ws["A1"]
    title.value = f"{settings.EVENT_FULL_NAME} — Registration List"
    title.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    title.fill = PatternFill("solid", fgColor=NAVY_DARK)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last_letter}2")
    sub = ws["A2"]
    generated = timezone.localtime(timezone.now()).strftime("%d %b %Y, %I:%M %p")
    bits = [f"{queryset.count()} record(s)", f"Generated {generated}"]
    if filter_summary:
        bits.insert(0, filter_summary)
    sub.value = "   •   ".join(bits)
    sub.font = Font(name="Calibri", size=10, italic=True, color=NAVY)
    sub.fill = PatternFill("solid", fgColor=CREAM)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 6  # spacer

    # ---- header row ------------------------------------------------------
    header_row = 4
    head_fill = PatternFill("solid", fgColor=NAVY)
    head_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    for idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=label)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[header_row].height = 26

    # ---- data rows -------------------------------------------------------
    body_font = Font(name="Calibri", size=10.5)
    row = header_row + 1

    for n, r in enumerate(queryset, start=1):
        values = [
            n,
            r.registration_id,
            r.full_name,
            r.phone,
            r.secondary_phone or "",
            r.whatsapp_number or "",
            r.email,
            r.last_class_attended,
            r.ssc_batch,
            r.ssc_passing_year or "",
            r.blood_group or "",
            r.get_tshirt_size_display() if r.tshirt_size else "",
            r.present_address or "",
            r.amount,
            r.get_payment_status_display(),
            r.transaction_id or "",
            _local(r.paid_at),
            "Yes" if r.is_checked_in else "No",
            _local(r.checked_in_at),
            _local(r.created_at),
        ]

        banded = (n % 2 == 0)
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = body_font
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 9, 10, 11, 12, 14, 18) else "left",
                vertical="center",
                wrap_text=(col == 13),
            )
            if banded:
                cell.fill = PatternFill("solid", fgColor=BAND)

        # amount as currency
        amount_cell = ws.cell(row=row, column=14)
        amount_cell.number_format = '#,##0'

        # colour-code the payment status
        status_cell = ws.cell(row=row, column=15)
        bg, fg = STATUS_FILL.get(r.payment_status, (BAND if banded else WHITE, NAVY))
        status_cell.fill = PatternFill("solid", fgColor=bg)
        status_cell.font = Font(name="Calibri", size=10.5, bold=True, color=fg)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

    last_data_row = row - 1

    # ---- totals ----------------------------------------------------------
    if last_data_row >= header_row + 1:
        ws.cell(row=row, column=13, value="TOTAL").font = Font(
            name="Calibri", size=11, bold=True, color=NAVY
        )
        ws.cell(row=row, column=13).alignment = Alignment(horizontal="right", vertical="center")
        total_cell = ws.cell(
            row=row, column=14,
            value=f"=SUM(N{header_row + 1}:N{last_data_row})",
        )
        total_cell.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        total_cell.number_format = '#,##0'
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in (13, 14):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=CREAM)
            ws.cell(row=row, column=col).border = BORDER

        ws.auto_filter.ref = f"A{header_row}:{last_letter}{last_data_row}"

    ws.freeze_panes = ws.cell(row=header_row + 1, column=4)

    # print setup — handy if anyone prints the list at the gate
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return wb
